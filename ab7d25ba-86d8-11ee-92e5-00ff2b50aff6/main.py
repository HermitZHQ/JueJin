# coding=utf-8
# ------------------Monitor：监控盯盘用（针对符合条件的标的，生成excel报告，以便进行查看）
# 监控策略，统一只用来做数据监控和表格生成，不做具体的买卖操作
# 在这个监控策略中，可能有多个监控策略（读取不同的ID列表，生成不同的报告表格，没有必要每个监控都单开策略，那样更浪费资源，除非确实有必须分离处理的理由）
# 各个监控的话，以名字区分，比如monitor1, monitor2, 用注释写清楚大概的监控逻辑即可

from __future__ import print_function, absolute_import
from gm.api import *
from threading import Timer

import time
import datetime
import math
import numpy as np
import pickle
import re

# excel库
import openpyxl
from openpyxl.utils import get_column_letter #导入合并单元格模块
from openpyxl.styles import Alignment # 导入对齐模块
from openpyxl.styles import Font # 导入字体模块
from openpyxl.styles import Border, Side # 导入边框线和颜色模块
from openpyxl.styles import PatternFill # 导入填充模块

# ！！警告！！复制这份代码的时候一定要注意修改下面的文件路径 + 策略模式 + 买入模式，其他不用改
# 这样就可以把使用策略在一份代码内进行维护了，虽然量大，但是封装好的话，问题不大
def StrategyM():
    pass

# 全局需要修改的变量，如果策略变化（比如买卖变化，策略本身变化）都应该调整下面的值
str_strategy = 'M'
log_path = 'c:\\TradeLogs\\Trade' + str_strategy + '.txt'
ids_path_a1 = 'c:\\TradeLogs\\IDs-' + str_strategy + '-A1.txt'
ids_path_a2 = 'c:\\TradeLogs\\IDs-' + str_strategy + '-A2.txt'
pos_info_path = 'c:\\TradeLogs\\Pos-' + str_strategy + '.npy'
statistics_info_path = 'c:\\TradeLogs\\Sta-' + str_strategy + '.npy'
buy_info_path = 'c:\\TradeLogs\\Buy-' + str_strategy + '.npy'

side_type = OrderSide_Buy # 设置买卖方向，买卖是不一样的，脚本切换后，需要修改
order_overtime = 3 # 设置的委托超时时间，超时后撤单，单位秒
sell_all_time = "13:35"

class BuyMode:
    def __init__(self):
        # 注意每个策略下面只能设定一种对应的买入模式（也就是只有一种可以激活到1）！！
        self.buy_all = 1 # 是否和策略B一样，直接从列表一次性全部买入，使用读出的配置数据就可以动态算出每只票的分配仓位
        self.buy_one = 0 # 买一只，并指定对应的价格
        self.buy_all_force = 0 # 一定要保证买入所有（对应数量非常大的买入，且有很多高价股，均分值无法覆盖高价股）

class OrderTypeBuy:
    def __init__(self):
        # 选择整体的订单交易模式，限价或者市价，只能激活一种！！
        self.Limit = 1
        self.Market = 0

class OrderTypeSell:
    def __init__(self):
        # 选择整体的订单交易模式，限价或者市价，只能激活一种！！
        self.Limit = 0
        self.Market = 1

class StrategyInfo:
    def __init__(self):
        # 指定策略对应的模式，策略A就对应A=1，一次只能激活一种策略！！
        self.A = 0
        self.AC = 0
        self.A1 = 0
        self.AA = 0
        self.B = 0
        self.B1 = 0
        self.BA = 0
        self.C = 0
        self.M = 1

        self.buy_mode = BuyMode()
        self.order_type_buy = OrderTypeBuy()
        self.order_type_sell = OrderTypeSell()

        self.slip_step = 2 # 滑点级别，1-5，要突破5的话，可以自己算？
        
class MonitorA1Info:
    def __init__(self):
        self.over_3_time = 0
        self.lowest_rate = 1.0
        self.highest_rate = -1.0
        self.lowest_rate_time = 0
        self.highest_rate_time = 0
        self.close_rate = 0

class LoadedIDsInfo:
    def __init__(self):
        self.buy_flag = 0 # 0->卖出，1->买入，2->??暂留空
        self.already_buy_in_flag = False # 已经被买入标记（用于order中更新，并自动收盘后更新ids文档）
        self.high_expected_flag = 0 # 策略A的高预期标记
        self.force_sell_flag = 0 # 无视其他条件直接卖出
        self.force_buy_flag = 0 # 无视其他条件直接买入
        self.buy_amount = 0 # 单个买入模式，可以支持设置买入资金，或者余额比例
        #self.buy_rate = 0 # 不记录buy_rate，因为rate会随着余额的变化而变化，我们只在第一次设置ids的时候，计算一次应该买入的值，[TODO]实际这个值还应该记录进入文件
        self.buy_with_rate = 1
        self.buy_with_time = "15:51"
        self.buy_with_price = 0
        self.buy_with_num = 0 # 设定的单次买入股数，买入后，再次刷新的话会再次买入指定数量，需要注意
        self.buy_with_num_need_handle = False # 是否需要处理买入股数（设定数量大于0时，把标记激活）
        self.buy_with_num_handled_flag = False # 是否已经处理完毕（已经调用了买入接口，把标记激活，用于后续逻辑判断）
        self.sell_with_rate = 1
        self.sell_with_time = "15:51"
        self.sell_with_price = 0
        self.sell_with_num = 0 # 指定卖出的股数（非全部卖出时使用）

        # 加入变量，控制tick的处理次数，主要为了处理AA这种数量太多的情况，如果每次tick都处理，会造成很大的延迟，导致refresh卡很久
        # 像B这种少的话，我们就每次处理就好
        self.tick_cur_count = 0
        self.tick_handle_frequence = 1

        # 多次上涨后买入，相关配置参数
        self.enable_multi_up_buy_flag = False # 多次上涨才买入的启用开关
        self.multi_up_time_line = "09:45" # 检测多次上涨的时间线
        self.highest_open_fpr = -1.0 # 记录时间线前最高的涨幅情况，针对今开价格
        self.multi_up_buy_flag = False # 多次上涨到目标比例后才买入（为了应对2次或更多到达目标后买入）
        self.multi_up_buy_need_check = True # 用于开启脚本后的检查，比如当下超过X%，就认为是需要处理（进一步等待下跌到X以下，加一次cur_count)
        self.multi_up_total_count = 1 # 总共需要检查的次数，虽然目前需求是1，但是直接做成多次的以后方便
        self.multi_up_cur_count = 0

class TargetInfo:
    def __init__(self):
        self.name = ""
        self.hold = 0
        self.price = 0
        self.first_record_flag = False
        self.pre_close = 0
        self.vwap = 0
        self.upper_limit = 0 # 涨停价
        self.lower_limit = 0
        self.suspended = False # 是否停牌
        self.sold_flag = False
        self.sold_price = 0 # 虚拟卖出时的记录价格，用于统计信息
        self.sold_mv = 0 # 虚拟卖出后锁定的市值
        self.fpr = -1 # 浮动盈亏缓存
        self.total_holding = 0 # 这里的总量并不从pos中取，而是从完成（部成）的order中记录，因为我发现完全用pos中的有时候数据更新不及时，比如部成的数量都已经显示有10000股了，且已经输出到日志中，但是紧接着马上取pos中的持仓，却还没有更新到该值，取出来可能是4，5000，虽然这个发生的几率很小，但是还是需要处理
        self.partial_holding = 0 # 记录部成的临时值
        self.fixed_buy_in_base_num = 0 # 强制买入所有目标下使用的变量，记录需要买入的base数量（手，最后需要乘100）

class MaxMinInfo:
    def __init__(self):
        self.max = -1.0
        self.min = 1.0
        self.time = None

class StatisticsInfo:
    def __init__(self):
        # 最高最低整体盈利记录
        self.lowest_total_fpr = 1.0
        self.lowest_total_fpr_time = None
        self.highest_total_fpr = -1.0
        self.highest_total_fpr_time = None
        self.cur_fpr = 0

        # limit ver为统计的虚拟止盈止损版本
        self.lowest_total_fpr_limit_ver = 1.0
        self.lowest_total_fpr_time_limit_ver = None
        self.highest_total_fpr_limit_ver = -1.0
        self.highest_total_fpr_time_limit_ver = None
        self.cur_fpr_limit_ver = 0

        # 各个卖出标的的当日最高和最低值记录
        # key:symbol, value: MaxMinFin()
        self.max_min_info_dict = {}
        

def log(msg):
    file_obj = open(log_path, 'a')

    nowtime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(time.time()))
    finalMsg = nowtime + ": " + msg + "\n"
    file_obj.write(finalMsg)

    # 我们也可以把所有日志打印到控制台查看
    print(finalMsg)

def refresh(context):
    context.ids = {}
    load_ids_a1(context)
    context.get_all_buy_price_flag = False
    context.get_all_sell_price_flag = False

    # 开始订阅目标，这里就比较麻烦了，无法快速输入
    # 统计买入和卖出的单独数量
    t = time.time()
    buy_num = 0
    handled_num = 0 # 用于计算初始化进度
    context.ids_buy = []
    context.ids_sell = []
    for k, v in context.ids.items():
        if (v.buy_flag == 1):
            if k not in context.ids_buy_target_info_dict.keys():
                context.ids_buy_target_info_dict[k] = TargetInfo()

            # 在初始化的时候就直接获取各项缓存值（反正后面也会获取，而且这里获取，会更集中）
            # 主要在这里我们可以过滤掉停牌的标的
            suspended = False
            if context.ids_buy_target_info_dict[k].pre_close == 0:
                info = get_instruments(symbols = k, skip_suspended = False, df = True)
                # empty情况一般就是ST的股票，直接先跳过不处理
                if info.empty:
                    print(f"[init][{k}]get cache info null, this should not happen......")
                    continue
                    
                # 最好不要直接使用df的值，很多奇怪的现象，比如下面的float相除，如果插入了df数据，结果是对的，但是小数点只有2位。。。。
                # 一定要先从数组里拿出来
                context.ids_buy_target_info_dict[k].pre_close = info.pre_close[0]
                context.ids_buy_target_info_dict[k].name = info.sec_name[0]
                context.ids_buy_target_info_dict[k].upper_limit = info.upper_limit[0]
                context.ids_buy_target_info_dict[k].suspended = True if (info.is_suspended[0] == 1) else False

            suspended = context.ids_buy_target_info_dict[k].suspended

            if not suspended:
                buy_num += 1
                context.ids_buy.append(k)
            else:
                del context.ids_buy_target_info_dict[k]
        else:
            context.ids_sell.append(k)
            if k not in context.ids_virtual_sell_target_info_dict.keys():
                context.ids_virtual_sell_target_info_dict[k] = TargetInfo()
            if k not in context.statistics.max_min_info_dict.keys():
                context.statistics.max_min_info_dict[k] = MaxMinInfo()
        handled_num += 1
        # print(f"初始化数据进度：[{round(handled_num / len(context.ids.items()) * 100, 3)}%]")
    # print(f"初始化总计耗时:[{time.time() - t:.4f}]s")
    context.buy_num = buy_num
    context.sell_num = len(context.ids) - buy_num

    # 检测一次是否有删除的id，在保存的ids_virtual_sell_target_info_dict中，应该去除（停牌等），否则无法正常统计
    del_keys = []
    for k in context.ids_virtual_sell_target_info_dict.keys():
        # 这里的条件不能是不存在就删，因为还有一种情况需要保留，就是sell列表为空的时候（否则删除后无法继续统计信息）
        # 也就是说，我们在sell为空时，并不删除缓存对象，只在sell列表存在，并只清除一部分的时候才删除（这时候我们是删除不需要，或者停牌之类的）
        if (k not in context.ids.keys()) and (len(context.ids_sell) > 0):
            # 不能在迭代的途中去删除目标中的值，会出现运行错误，先记录下来，然后再单独删除
            del_keys.append(k)
    for k in del_keys:
        del context.ids_virtual_sell_target_info_dict[k]

    # ------看是否需要重新统计卖出票的总市值（注意是卖出票，不含买入）
    # 即便设置不是-1，我们也应该检查一次（我之前就误操作，前一天下午开过脚本，忘记重置-1，结果load的都是昨天的sell，导致今天的脚本完全无法正常运行）
    if context.calculate_total_market_value_flag:
        context.total_market_value_for_all_sell = 0
        for symb in context.ids_sell:
            pos = context.account().position(symbol = symb, side = OrderSide_Buy)
            amount = (pos.available_now * pos.vwap) if pos else 0
            context.total_market_value_for_all_sell += amount
        
        if context.total_market_value_for_all_sell > 0:
            over_write_mv(context.total_market_value_for_all_sell)
            save_sell_position_info_with_init(context)
            save_statistics_info(context)
            load_buy_info(context)

    # 频率, 支持 ‘tick’, ‘60s’, ‘300s’, ‘900s’ 等, 默认’1d’
    # 60s以上都是走on_bar，只有tick走的on_tick，一般来说交易肯定是要走tick的
    subscribe(symbols=context.ids.keys(), frequency='tick', count=1, unsubscribe_previous=True)
    # 用1s频率的bar函数来更新总盈亏，否则tick中循环用current很慢
    #subscribe(symbols='SZSE.000001', frequency='15s', count=1, unsubscribe_previous=False)
    #log(f'已订阅目标：{context.ids.keys()} 总数量：{len(context.ids)} 买入数量：{buy_num} 卖出数量：{len(context.ids) - buy_num}')
    #log(f'已订阅：{context.symbols} 数量：{len(context.symbols)}')    

    # 手动清空buy info，里面记录的买入价格，文件太大需要清空，则修改控制参数后，点击刷新
    if (context.clear_buy_info_flag):
        context.buy_pos_dict = {}
        save_buy_info(context)


def refresh_statistics_info(context):
    # 刷新后显示一下目前最新的最高值和最低值
    print(f"最高：{round(context.statistics.highest_total_fpr * 100, 3)}% 出现时间：{context.statistics.highest_total_fpr_time}")
    print(f"最低：{round(context.statistics.lowest_total_fpr * 100, 3)}% 出现时间：{context.statistics.lowest_total_fpr_time}")
    print(f"当前盈利：{round(context.statistics.cur_fpr * 100, 3)}%\n")

    print(f"[止]最高：{round(context.statistics.highest_total_fpr_limit_ver * 100, 3)}% 出现时间：{context.statistics.highest_total_fpr_time_limit_ver}")
    print(f"[止]最低：{round(context.statistics.lowest_total_fpr_limit_ver * 100, 3)}% 出现时间：{context.statistics.lowest_total_fpr_time_limit_ver}")
    print(f"[止]当前盈利：{round(context.statistics.cur_fpr_limit_ver * 100, 3)}%\n")


# 统一所有策略的配置文件读取，不然每个策略都要重写没有必要，想办法加入必要的配置就好
def load_ids_a1(context):
    # 看股票代码就能分辨，上证股票是在上海交易所上市的股票，股票代码以600、601、603开头，科创板（在上海交易所上市）股票代码以688开头
    # 深市股票是在深市交易所上市的股票，股票代码以000、002、003开头，创业板（在深圳交易所上市）股票代码以300开头。
    # 
    # 简化使用的话，就是6开头就都是上交所的，其他都是深交所（当然这种情况没有处理错误的存在）
    # SHSE(上交所), SZSE(深交所)

    # ！！！这里读取的配置文件，要根据策略改变
    file_obj = open(ids_path_a1, 'r')
    lines = file_obj.readlines()

    buy_flag = True
    for line in lines:
        should_add = True
        str_tmp = line.strip() # 去掉换行符
        original_id = str_tmp

        # 属于一个技巧性地读取，首先Buy在配置文件的上面，Sell在下面
        # 所以上面都是true，当读取到--------Sell标志行的时候，转为False，这样下面的都是False
        if (str_tmp.find('--------Sell') != -1):
            buy_flag = False
        first3 = str_tmp[:3]

        if (first3 == '600'):
            str_tmp = 'SHSE.' + str_tmp[:6]
        elif (first3 == '601'):
            str_tmp = 'SHSE.' + str_tmp[:6]
        elif (first3 == '603'):
            str_tmp = 'SHSE.' + str_tmp[:6]
        elif (first3 == '605'):
            str_tmp = 'SHSE.' + str_tmp[:6]
        elif (first3 == '688'):
            str_tmp = 'SHSE.' + str_tmp[:6]
        elif (first3 == '689'):
            str_tmp = 'SHSE.' + str_tmp[:6]
        elif (first3 == '000'):
            str_tmp = 'SZSE.' + str_tmp[:6]
        elif (first3 == '001'):
            str_tmp = 'SZSE.' + str_tmp[:6]
        elif (first3 == '002'):
            str_tmp = 'SZSE.' + str_tmp[:6]
        elif (first3 == '003'):
            str_tmp = 'SZSE.' + str_tmp[:6]
        elif (first3 == '300'):
            str_tmp = 'SZSE.' + str_tmp[:6]
        elif (first3 == '301'):
            str_tmp = 'SZSE.' + str_tmp[:6]
        elif (first3 == 'fsa'):
            should_add = False
            context.force_sell_all_flag = True
        elif (first3 == 'mv:'):
            should_add = False
            mv = float(str_tmp[3:]) if (str_tmp[3:] != '') else 0.0
            if -1 == mv: # 初始化的情况（应该只有每天第一次启动脚本的时候执行这里）
                context.calculate_total_market_value_flag = True
                log(f"did not find valid market value, need re-calculate")
            else:
                log(f"find valid total market value:{mv}")
                context.calculate_total_market_value_flag = False
                context.total_market_value_for_all_sell = mv
                load_sell_position_info_with_init(context)
                load_statistics_info(context)
                load_buy_info(context)
        else:
            should_add = False
            print(f'读取IDs配置错误：{str_tmp}')

        # 开始我固定了一些特殊标记的位置，比如高预期一定要是最后一位的h，卖出时倒数第二位，其实没有必要
        # 我们判断长度大于6（股票代码）后，直接把后几位的字符都拿出来保存，然后判断有没有就可以了
        flags = []
        if should_add and len(original_id) > 6:
            for i in range(0, len(original_id) - 6):
                flags.append(original_id[6 + i])

            #print(flags)

        # 判断是否为High Expect（高预期）股票
        high_expect = True if 'h' in flags else False
        # 判断是否要无视条件直接卖出
        force_sell = True if 's' in flags else False
        # 判断是否要无视条件直接买入
        force_buy = True if 'b' in flags else False

        buy_amount = 0
        buy_rate = 0
        # 判定是否设置了单独买入的数量或者比例
        if ('#' in flags):
            if '%' in flags:
                buy_rate = float(re.findall(r"#(.*?)%#", original_id)[0]) / 100.0
            else:
                buy_amount = float(re.findall(r"#(.*?)#", original_id)[0]) * 10000.0

            #print(f"buy amount:{buy_amount}, buy_rate:{buy_rate}")

        left_cash = context.account().cash['available'] # 余额
        left_cash = float('%.2f' % left_cash)
        
        if buy_rate > 0:
            buy_amount = left_cash * buy_rate
        # if buy_amount > left_cash:
        #     log(f"[Error]设置的买入金额:{buy_amount}超过了余额:{left_cash}，自动将数值降低到目前余额")
        #     buy_amount = left_cash

        # 判定是否设定了卖出（买入）的比例（R）或者时间（T），比如设定卖出比例就是SR1SR，表示上涨超过1%卖出
        # ST10:35ST，标识到达时间10点35分后就卖出，买入的话就是把S改为B，记得一定要大写，和小写的强制买入卖出标记区别开来
        sell_with_rate = 1
        sell_with_time = "17:31" # 初始化到收盘后
        sell_with_price = 0
        sell_with_num = 0
        buy_with_rate = 1
        buy_with_time = "17:31"
        buy_with_price = 0
        buy_with_num = 0

        if ('S' in flags) and ('R' in flags):
            sell_with_rate = float(re.findall(r"SR(.*?)SR", original_id)[0]) / 100.0
        if ('S' in flags) and ('T' in flags):
            sell_with_time = re.findall(r"ST(.*?)ST", original_id)[0]
        if ('S' in flags) and ('P' in flags):
            sell_with_price = float(re.findall(r"SP(.*?)SP", original_id)[0])
        if ('S' in flags) and ('N' in flags):
            sell_with_num = float(re.findall(r"SN(.*?)SN", original_id)[0])
        if ('B' in flags) and ('R' in flags):
            buy_with_rate = float(re.findall(r"BR(.*?)BR", original_id)[0]) / 100.0
        if ('B' in flags) and ('T' in flags):
            buy_with_time = re.findall(r"BT(.*?)BT", original_id)[0]
        if ('B' in flags) and ('P' in flags):
            buy_with_price = float(re.findall(r"BP(.*?)BP", original_id)[0])
        if ('B' in flags) and ('N' in flags):
            buy_with_num = float(re.findall(r"BN(.*?)BN", original_id)[0])

        # 取数组最后n位（-1，-2......）
        #print(f"test -1::::: {str_tmp[-1:]}")sss
        
        # test info, open it when you need debug
        #print(f'high expect:{high_expect} force sell:{force_sell} force buy:{force_buy} buy_amount:{buy_amount} buy_rate:{buy_rate} sell_with_up:{sell_with_rate} sell_with_time:{sell_with_time} buy_with_up:{buy_with_rate} buy_with_time:{buy_with_time}')
        if (should_add):
            if str_tmp in context.ids.keys():
                context.ids[str_tmp].buy_flag = 1 if buy_flag else 0
                context.ids[str_tmp].high_expected_flag = high_expect
                context.ids[str_tmp].force_sell_flag = force_sell
                context.ids[str_tmp].force_buy_flag = force_buy

                context.ids[str_tmp].buy_with_rate = buy_with_rate
                context.ids[str_tmp].buy_with_time = buy_with_time
                context.ids[str_tmp].buy_with_price = buy_with_price
                context.ids[str_tmp].buy_with_num = buy_with_num
                if buy_with_num > 0:
                    context.ids[str_tmp].buy_with_num_need_handle = True
                    context.ids[str_tmp].buy_with_num_handled_flag = False
                else:
                    context.ids[str_tmp].buy_with_num_need_handle = False
                context.ids[str_tmp].sell_with_rate = sell_with_rate
                context.ids[str_tmp].sell_with_time = sell_with_time
                context.ids[str_tmp].sell_with_price = sell_with_price
                context.ids[str_tmp].sell_with_num = sell_with_num

                # 在ids已经存在的情况下，我们不刷新买入数额，数额只在key不存在的时候记录
                # 否则有些问题，比如我们设置的10%，但是余额是变化的，如果变多了，我们就会去买入更多，就不正常了
                # 没有设置买入百分比的话，我们可以重置买入数量
                if buy_rate == 0:
                    context.ids[str_tmp].buy_amount = buy_amount
            else:
                context.ids[str_tmp] = LoadedIDsInfo()
                context.ids_a1[str_tmp] = LoadedIDsInfo()
                context.ids[str_tmp].buy_flag = 1 if buy_flag else 0
                context.ids[str_tmp].high_expected_flag = high_expect
                context.ids[str_tmp].force_sell_flag = force_sell
                context.ids[str_tmp].force_buy_flag = force_buy

                context.ids[str_tmp].buy_with_rate = buy_with_rate
                context.ids[str_tmp].buy_with_time = buy_with_time
                context.ids[str_tmp].buy_with_price = buy_with_price
                context.ids[str_tmp].buy_with_num = buy_with_num
                if buy_with_num > 0:
                    context.ids[str_tmp].buy_with_num_need_handle = True
                    context.ids[str_tmp].buy_with_num_handled_flag = False
                else:
                    context.ids[str_tmp].buy_with_num_need_handle = False
                context.ids[str_tmp].sell_with_rate = sell_with_rate
                context.ids[str_tmp].sell_with_time = sell_with_time
                context.ids[str_tmp].sell_with_price = sell_with_price
                context.ids[str_tmp].sell_with_num = sell_with_num

                # 只在初始化时，计算一次应该买入的量，否则后面余额发生变化就不对了
                # [TODO]应该保存文件更稳
                context.ids[str_tmp].buy_amount = buy_amount
            context.ids_a1[str_tmp] = context.ids[str_tmp]

    #print(f"below is context ids info:\n{context.ids}")


def over_write_mv(new_val):
    pass

def over_write_force_sell_all_flag(new_val):
    pass

def auto_generate_sell_list_with_ids_file(context):
    pass


def save_sell_position_info_with_init(context):
    context.sell_pos_dict.clear()
    for symb in context.ids_sell:
        pos = context.account().position(symbol = symb, side = OrderSide_Buy)
        curHolding = pos.available_now if pos else 0
        # 如果持仓为0，则不要进行记录，比如有时候ID列表里面设置了错误的ID,该ID并没有持仓，这会造成后续的统计错误
        if curHolding == 0:
            continue
        # 数据说明：持仓量，是否已经卖出，卖出时的均价（用于计算整体盈利）
        context.sell_pos_dict[symb] = [curHolding, False, 0.0]
    
    np.save(pos_info_path, context.sell_pos_dict)

# 更新信息只在全部卖出后，才记录卖出的成交均价，持仓量，我们只在当天第一次开启脚本的时候获取就可以了
def update_sell_position_info(context, symbol, sell_flag, sell_price):
    if symbol in context.sell_pos_dict.keys():
        context.sell_pos_dict[symbol][1] = sell_flag
        context.sell_pos_dict[symbol][2] = sell_price
        print(f"update sell pos info sym:{symbol} sell_flag:{sell_flag}")
        np.save(pos_info_path, context.sell_pos_dict)

    # 卖出后，清理buy_pos_dict中保存的值
    # 和2点全部清理进行互补，万一2点之前就开始买入了，这里也能有较大的作用，只要不是通过手动卖出的
    if symbol in context.buy_pos_dict.keys():
        del context.buy_pos_dict[symbol]

    for k,v in context.sell_pos_dict.items():
        if not v[1]:
            return # 只要还有没卖出的，我们就不统计精确的整体盈利值

    cur_market_value = 0
    for k,v in context.sell_pos_dict.items():
        if not v[1]:
            # mv = float(current(k)[0]['price']) * v[0]
            # cur_market_value += mv
            log(f"should never happen...., just in case")
        else:
            cur_market_value += (v[0] * v[2])

    # 现在策略AB应该是可以共用一套了，因为不管是否存在买入和卖出的混合，我们有了当天的文件缓存后，都可以计算正确的整体盈利了
    # 之前的话，因为没有文件缓存，所以逻辑上有点混乱，现在是很清晰的
    total_float_profit_rate = (cur_market_value - context.total_market_value_for_all_sell) / context.total_market_value_for_all_sell if context.total_market_value_for_all_sell != 0 else 0
    log(f"[statistics]全部卖出后的精确统计值(不加税费)，整体盈利为：{round(total_float_profit_rate * 100, 3)}%")


def load_sell_position_info_with_init(context):
    context.sell_pos_dict = np.load(pos_info_path, allow_pickle=True) # 输出即为Dict类型
    context.sell_pos_dict = dict(context.sell_pos_dict.tolist())
    #print("--------------------start load sell pos info")
    #print(context.sell_pos_dict)
    #print("--------------------end load sell pos info\n")

    for k,v in context.sell_pos_dict.items():
        if (k in context.ids.keys()) and (context.ids[k].buy_flag == 0) and (k not in context.ids_virtual_sell_target_info_dict.keys()):
            context.ids_virtual_sell_target_info_dict[k] = TargetInfo()

#change mark2
def save_buy_info(context):
    np.save(buy_info_path, context.buy_pos_dict)

def update_buy_info(context, symbol, price):
    if not context.buy_pos_dict:
        return
    log(f"[statistics]{symbol} 更新记录买入价格：{price}")
    context.buy_pos_dict[symbol] = price
    np.save(buy_info_path, context.buy_pos_dict)

def load_buy_info(context):
    context.buy_pos_dict = np.load(buy_info_path, allow_pickle=True) # 输出即为Dict类型
    context.buy_pos_dict = dict(context.buy_pos_dict.tolist())
    #print("--------------------start load buy pos info")
    #print(context.buy_pos_dict)
    #print("--------------------end load buy pos info\n")


def save_statistics_info(context):
    f = open(statistics_info_path, 'wb')
    pickle.dump(context.statistics, f, pickle.HIGHEST_PROTOCOL)


def load_statistics_info(context):
    f = open(statistics_info_path, 'rb')
    context.statistics = pickle.load(f) # 输出即为Dict类型
    #print("--------------------start load statistics info")
    #print(context.statistics.max_min_info_dict)
    #print("--------------------end load statistics info\n")


def on_parameter(context, parameter):
    #print(parameter)
    if (parameter.key == 'Buy_In_Fall_Rate'):
        print(f'Buy_In_Fall_Rate 参数已经调整为：{parameter.value}')
        context.Buy_In_Fall_Rate = parameter.value
    if (parameter.key == 'Buy_In_Up_Rate'):
        print(f'Buy_In_Up_Rate 参数已经调整为：{parameter.value}')
        context.Buy_In_Up_Rate = parameter.value
    elif (parameter.key == 'Buy_In_Limit_Num'):
        print(f'Buy_In_Limit_Num 参数已经调整为：{parameter.value}')
        context.Buy_In_Limit_Num = parameter.value
    elif (parameter.key == 'Use_Close'):
        print(f'Use_Close 参数已经调整为：{parameter.value}')
        context.Use_Close = parameter.value
    elif (parameter.key == 'Sell_Increase_Rate'):
        print(f'Sell_Increase_Rate 参数已经调整为：{parameter.value}')
        context.Sell_Increase_Rate = parameter.value
    elif (parameter.key == 'Sell_All_Increase_Rate'):
        print(f'Sell_All_Increase_Rate 参数已经调整为：{parameter.value}')
        context.Sell_All_Increase_Rate = parameter.value
    elif (parameter.key == 'Sell_All_Loss_Rate'):
        print(f'Sell_All_Loss_Rate 参数已经调整为：{parameter.value}')
        context.Sell_All_Loss_Rate = parameter.value
    elif (parameter.key == 'Sell_Loss_Limit'):
        print(f'Sell_Loss_Limit 参数已经调整为：{parameter.value}')
        context.Sell_Loss_Limit = parameter.value
    elif (parameter.key == 'clear_buy_info_flag'):
        print(f'clear_buy_info_flag 参数已经调整为：{parameter.value}')
        context.clear_buy_info_flag = parameter.value
    elif (parameter.key == 'test_info'):
        print(f'test_info 参数已经调整为：{parameter.value}')
        context.test_info = parameter.value
        context.ids_a1_manual_generate_excel_flag = True
        refresh_statistics_info(context)
        if context.test_info == 99:
            # over_write_mv(-1) # 这里重置为0问题不大，因为市值一般都是两位小数的float，比如1234.68，目前感觉是不会替换到有效id数据的
            # over_write_force_sell_all_flag('') # 重置强制卖出标记，避免忘记后，第二天被直接全卖
            # auto_generate_sell_list_with_ids_file(context)
            output_final_statistics(context)
    elif (parameter.key == 'Refresh'):
        log("重新载入ids，重新订阅！")
        refresh(context)


# 策略中必须有init方法
def init(context):
    # 订阅行情数量（-1表示全部订阅）
    # context.max_count = 5000
    # all_stock = get_instruments(exchanges='SHSE, SZSE', sec_types=1, fields='symbol, delisted_date', df=True)
    # subscribe(symbols=all_stock['symbol'].tolist()[:context.max_count], frequency='60s', count=5)
    # print('订阅标的数量', len(context.symbols))
    # return

    # 初始化一些全局用变量--------
    # 保存策略信息以及买入方式到全局变量中
    context.strategy_info = StrategyInfo()

    log("\n\n\n\n\n\n\n\n策略脚本初始化开始--------")
    context.test_info = 0
    context.record_sell_time = 0 # 用于记录整体卖出时的时间和瞬间整体盈利值
    context.record_sell_tfpr = 0
    add_parameter(key='test_info', value=context.test_info, min=-1, max=1, name='TestInfo', intro='', group='3', readonly=False)
    #print(f'{context.now.strftime("%H:%M:%S")}')
    context.record_time = 0 # 用于记录时间差，基本用法就是初始化 = time.time()，然后计算时间差用time.time() - 上次记录的时间，可以在下面搜索用法
    context.tick_count_for_statistics = 0 # 记录经过的tick次数，有些函数不需要每次tick都调用，会造成性能损耗，用这个变量进行控制
    context.tick_count_for_sell = 0 # 用于控制sell函数中一些消耗较大的代码块的运行频率
    context.tick_count_limit_rate_for_sell = 0.2 # 具体的控制百分比（针对context.ids的长度），比如有1000标的，那么0.2就是20%，200只标的经过后，再运行一次
    context.tick_count_for_sell_ok = False # 记录当前是否满足频率，避免每次都要和rate做乘法后比较
    context.tick_count_for_buy = 0
    context.tick_count_limit_rate_for_buy = 0.2
    context.tick_count_for_buy_ok = False 
    context.get_all_sell_price_flag = False
    context.get_all_buy_price_flag = False

    if context.strategy_info.order_type_buy.Limit == 1:
        context.order_type_buy = OrderType_Limit
    elif context.strategy_info.order_type_buy.Market == 1:
        context.order_type_buy = OrderType_Market

    if context.strategy_info.order_type_sell.Limit == 1:
        context.order_type_sell = OrderType_Limit
    elif context.strategy_info.order_type_sell.Market == 1:
        context.order_type_sell = OrderType_Market

    context.start_operation_time = "09:30"
    if context.strategy_info.A1 == 1:
        context.start_operation_time = "09:30"
    elif context.strategy_info.A == 1:
        context.start_operation_time = "09:30"

    # 整体数据为ditc，key=symbol，value = LoadedIDsInfo()
    context.ids = {}
    # 强制整体卖出（配置文件）标记
    context.force_sell_all_flag = False
    # 记录已买入的key（实际执行了买入接口的）
    context.already_buy_in_keys = []
    # 记录买入后的订单信息，防止重复买入
    context.client_order = {}
    # 记录最高的整理盈利情况以及当时的时间（该值只在连续运行脚本的情况下有效，统一在下午的2点59分输出到日志一次）
    context.statistics = StatisticsInfo()
    # 手动产生statistics info的初始文件（修改结构体后，之前的文件无法使用）
    if 0:
        context.statistics.highest_total_fpr = 0.3
        context.statistics.highest_total_fpr_time = context.now
        context.statistics.lowest_total_fpr = -0.05
        context.statistics.lowest_total_fpr_time = context.now
        context.statistics.cur_fpr = 0.123

        context.statistics.highest_total_fpr_limit_ver = 0.3
        context.statistics.highest_total_fpr_time_limit_ver = context.now
        context.statistics.lowest_total_fpr_limit_ver = -0.05
        context.statistics.lowest_total_fpr_time_limit_ver = context.now
        context.statistics.cur_fpr_limit_ver = 0.123
        print("start save statistics info................................")
        save_statistics_info(context)

    # 是否达到全局设定的盈利目标，初始化为false，但是后续条件一旦触发后，不再改回false
    # 因为操作那几秒，没法预估情况，有可能下降到1.99%，或者是先卖出了盈利高的，剩下的继续计算总盈利就不正确了
    # 所以一旦激活就直接全部卖出，忽略过程中的变化
    context.sell_with_total_float_profit_flag = False
    context.sell_with_total_float_loss_flag = False
    context.sell_with_time_limit_flag = False
    # 计算总市值，在开启脚本，且mv:0时统计一次，这样好计算全局的1.17%涨幅，关键要想办法自动重置到mv:0
    # 我目前想法是2点59分的时候，去把这个值覆写到mv:0
    context.calculate_total_market_value_flag = False
    context.total_market_value_for_all_sell = 0
    # 记录卖出持仓信息dict，这样才能保证全部卖出后，或者重新启动脚本时，能够拿到正常值    
    context.sell_pos_dict = {}
    
    # 记录买入的持仓信息，解决连续买入和卖出同一标的导致的成本上升或下降问题
    # 直接使用记录的成本值来替代持仓均价vwap
    # 这个数据的重置时间不是3点了，3点重置的是卖出相关数据，买入的话我们在1点35分，也就是设置的整体卖出之后重置它
    # 重置的时候有可能没有卖完，使用要判断使用的数据是否正常，比如为0的话，肯定就还是要用vwap替代
    # 暂时没有想到其他问题了
    context.buy_pos_dict = {} # buy info，需要手动清空
    context.clear_buy_info_flag = 0
    add_parameter(key='clear_buy_info_flag', value=context.clear_buy_info_flag, min=-1, max=1, name='是否清空BuyInfo', intro='', group='1', readonly=False)
    
    # 我们更新market value不能太频繁，否则就会出现如果有50只票，50个tick，每个都需要两秒，相当于100秒才能循环一次
    # 我开始思考过用时间差，比如控制1秒更新一次，其实也不行，因为数量上去以后，更新一次就要1秒多2秒，相当于还是一直更新
    # 目前看来靠谱的办法，就只有缓存上一帧的所有tick price，这样速度肯定最快，不用去服务器查询（current）
    # 只是这种方式，在第一次拿到所有有效值之前，肯定有很多0的价格，需要特殊处理，否则可能会激活整体亏损？（虽然暂时没有加这个）
    context.tfpr = 0.0 # total float profit rate
    # 该dict用于临时缓存所有标的的价格以及持仓成本，为了移除current查询价格带来的缓慢问题
    # 数据格式：key->symbol, value->TargetInfo()类
    # 该dict是用于虚拟止盈止损版本的信息统计，有些数据不能乱用，比如sold_mv，不是真实的卖出mv，需要注意！！
    context.ids_virtual_sell_target_info_dict = {}
    context.ids_buy_target_info_dict = {}
    # 监控脚本中，各个不同监控还需要有一套自己的保存id
    # A1相关保存数据
    context.ids_a1 = {}
    context.ids_a1_info_dict = {} # 以超过3%后的时间，进行存储的dict，后续应该需要进行排序，看如何处理
    context.ids_a1_time_sorted_arr = []
    context.ids_a1_excel_generated_flag = False # 全局标识，用于3点30输出一次excel文件结果
    context.ids_a1_manual_generate_excel_flag = False # 可用于中途手动生成excel文件，收盘时临时用当下的涨幅填充
    # 后续扩展监控保存数据，下面的x只是示例，并没有使用
    context.ids_x = {}

    # 手动产生buy_info的初始文件
    # save_buy_info(context)

    # 初始化动态加载的id文件--------
    load_ids_a1(context)

    # 初始化动态参数--------
    # 买入参数
    context.Buy_In_Fall_Rate = 0.01
    add_parameter(key='Buy_In_Fall_Rate', value=context.Buy_In_Fall_Rate, min=-1, max=1, name='买入时下跌比例', intro='', group='1', readonly=False)
    context.Buy_In_Up_Rate = 0.03
    add_parameter(key='Buy_In_Up_Rate', value=context.Buy_In_Up_Rate, min=-1, max=1, name='买入时上涨比例', intro='', group='1', readonly=False)

    context.Buy_In_Limit_Num = 5
    add_parameter(key='Buy_In_Limit_Num', value=context.Buy_In_Limit_Num, min=0, max=1, name='策略买入数量', intro='', group='1', readonly=False)
    context.Use_Close = 1
    add_parameter(key='Use_Close', value=context.Use_Close, min=0, max=1, name='是否使用昨收', intro='', group='1', readonly=False)
    # 卖出参数
    context.Sell_Increase_Rate = 0.2
    if context.strategy_info.B == 1:
        context.Sell_Increase_Rate = 0.07
    elif context.strategy_info.BA == 1:
        context.Sell_Increase_Rate = 0.07
    add_parameter(key='Sell_Increase_Rate', value=context.Sell_Increase_Rate, min=-1, max=1, name='卖出时涨幅比例', intro='', group='2', readonly=False)

    context.Sell_Loss_Limit = -0.2
    if context.strategy_info.B == 1:
        context.Sell_Loss_Limit = -0.035
    elif context.strategy_info.BA == 1:
        context.Sell_Loss_Limit = -0.035
    elif context.strategy_info.B1 == 1:
        context.Sell_Loss_Limit = -0.065
    add_parameter(key='Sell_Loss_Limit', value=context.Sell_Loss_Limit, min=-1, max=1, name='止损比例', intro='', group='2', readonly=False)

    context.Sell_All_Increase_Rate = 0.2
    context.sell_all_chase_raise_flag = False
    context.sell_all_chase_highest_record = -1.0
    context.sell_all_lower_limit = 0.0066
    if context.strategy_info.B == 1:
        context.Sell_All_Increase_Rate = 0.0066
    elif context.strategy_info.BA == 1:
        context.Sell_All_Increase_Rate = 0.0045
    elif context.strategy_info.A == 1:
        context.Sell_All_Increase_Rate = 0.0113
        
    context.Sell_All_Loss_Rate = -0.2
    if context.strategy_info.B == 1:
        context.Sell_All_Loss_Rate = -0.016
    elif context.strategy_info.BA == 1:
        context.Sell_All_Loss_Rate = -0.016
    add_parameter(key='Sell_All_Increase_Rate', value=context.Sell_All_Increase_Rate, min=-1, max=1, name='整体涨幅', intro='', group='2', readonly=False)
    add_parameter(key='Sell_All_Loss_Rate', value=context.Sell_All_Loss_Rate, min=-1, max=1, name='整体止损', intro='', group='2', readonly=False)
    # 取巧参数（参数赋值没有实际意义，只是用来重载ids配置文件），该参数变化后，重新load ids，重新订阅最新加载的ids
    context.Refresh = 0.01
    add_parameter(key='Refresh', value=context.Refresh, min=-1, max=1, name='刷新', intro='', group='4', readonly=False)

    # 开始订阅目标，这里就比较麻烦了，无法快速输入
    # 统计买入和卖出的单独数量
    t = time.time()
    buy_num = 0
    handled_num = 0 # 用于显示处理进度
    context.ids_buy = []
    context.ids_sell = []
    for k, v in context.ids.items():
        if (v.buy_flag == 1):
            print(f"start get info of [{k}]")
            context.ids_buy_target_info_dict[k] = TargetInfo()

            # 在初始化的时候就直接获取各项缓存值（反正后面也会获取，而且这里获取，会更集中）
            # 主要在这里我们可以过滤掉停牌的标的
            suspended = False
            if context.ids_buy_target_info_dict[k].pre_close == 0:
                info = get_instruments(symbols = k, skip_suspended = False, df = True)
                # empty情况一般就是ST的股票，直接先跳过不处理
                if info.empty:
                    print(f"[init][{k}]get cache info null, this should not happen......")
                    continue
                    
                # 最好不要直接使用df的值，很多奇怪的现象，比如下面的float相除，如果插入了df数据，结果是对的，但是小数点只有2位。。。。
                # 一定要先从数组里拿出来
                context.ids_buy_target_info_dict[k].pre_close = info.pre_close[0]
                context.ids_buy_target_info_dict[k].name = info.sec_name[0]
                context.ids_buy_target_info_dict[k].upper_limit = info.upper_limit[0]
                context.ids_buy_target_info_dict[k].suspended = True if (info.is_suspended[0] == 1) else False

            suspended = context.ids_buy_target_info_dict[k].suspended

            # 没有停牌的情况，才处理
            if not suspended:
                buy_num += 1
                context.ids_buy.append(k)
            else:
                del context.ids_buy_target_info_dict[k]
        else:
            context.ids_sell.append(k)
            context.ids_virtual_sell_target_info_dict[k] = TargetInfo()
            context.statistics.max_min_info_dict[k] = MaxMinInfo()
        handled_num += 1
        print(f"初始化数据进度：[{k}] [{round(handled_num / len(context.ids.items()) * 100, 3)}%]")
    print(f"初始化总计耗时:[{time.time() - t}]s")
    context.buy_num = buy_num
    context.sell_num = len(context.ids) - buy_num

    # 检测一次是否有删除的id，在保存的ids_virtual_sell_target_info_dict中，应该去除（停牌等），否则无法正常统计
    del_keys = []
    for k in context.ids_virtual_sell_target_info_dict.keys():
        # 这里的条件不能是不存在就删，因为还有一种情况需要保留，就是sell列表为空的时候（否则删除后无法继续统计信息）
        # 也就是说，我们在sell为空时，并不删除缓存对象，只在sell列表存在，并只清除一部分的时候才删除（这时候我们是删除不需要，或者停牌之类的）
        if (k not in context.ids.keys()) and (len(context.ids_sell) > 0):
            # 不能在迭代的途中去删除目标中的值，会出现运行错误，先记录下来，然后再单独删除
            del_keys.append(k)
    for k in del_keys:
        del context.ids_virtual_sell_target_info_dict[k]

    # ------看是否需要重新统计卖出票的总市值（注意是卖出票，不含买入）
    # 即便设置不是-1，我们也应该检查一次（我之前就误操作，前一天下午开过脚本，忘记重置-1，结果load的都是昨天的sell，导致今天的脚本完全无法正常运行）
    if context.calculate_total_market_value_flag:
        context.total_market_value_for_all_sell = 0
        for symb in context.ids_sell:
            pos = context.account().position(symbol = symb, side = OrderSide_Buy)
            amount = (pos.available_now * pos.vwap) if pos else 0
            context.total_market_value_for_all_sell += amount
        
        if context.total_market_value_for_all_sell > 0:
            over_write_mv(context.total_market_value_for_all_sell)
            save_sell_position_info_with_init(context)
            save_statistics_info(context)
            load_buy_info(context)

    # 手动生成buy_info，现在流程还不太稳定，有时候会误删，这里是手动生成
    # 注意是从buy还是sell的list生成，可能需要改动
    if 0:
        for symb in context.ids_sell:
            context.buy_pos_dict[symb] = 0
            pos = context.account().position(symbol = symb, side = OrderSide_Buy)
            update_buy_info(context, symb, pos.vwap)


    # --------订阅接口
    # 频率, 支持 ‘tick’, ‘60s’, ‘300s’, ‘900s’ 等, 默认’1d’
    # 60s以上都是走on_bar，只有tick走的on_tick，一般来说交易肯定是要走tick的
    subscribe(symbols=context.ids.keys(), frequency='tick', count=1, unsubscribe_previous=True)
    # 用1s频率的bar函数来更新总盈亏，否则tick中循环用current很慢
    #subscribe(symbols='SZSE.000001', frequency='1s', count=1, unsubscribe_previous=False)
    #log(f'已订阅目标：{context.ids.keys()} 总数量：{len(context.ids)} 买入数量：{buy_num} 卖出数量：{len(context.ids) - buy_num}')
    #log(f'已订阅：{context.symbols} 数量：{len(context.symbols)}')

    # 查询所有A股代码，暂时用不到，记录在这里
    #id_list = get_instruments(exchanges='SZSE,SHSE', sec_types=1, fields='symbol',df=1)['symbol'].tolist()
    #print(f"A股总数量：{len(id_list)}")

    print(context.now)
    #history_data = history(symbol='SHSE.600021', frequency='1d', start_time='2023-05-16',  end_time=datetime.datetime.now().strftime('%Y-%m-%d'), fields='symbol, open, close, low, high, amount', adjust=ADJUST_PREV, df= True)
    #print(history_data)
    #print(history_data.symbol[0])
    #print(history_data.amount)
    
    leftCash = context.account().cash['available'] # 余额
    leftCash = float('%.2f' % leftCash)
    totalCash = context.account().cash['nav'] # 总资金
    totalCash = float('%.2f' % totalCash)
    print(f'当前总市值：{totalCash} 资金余额：{leftCash}')
    
    totalBuyNum = len(context.ids_buy)
    curHoldTypeNum = len(context.account().positions(symbol='', side=None))
    print(f'策略设置买入品种数量：{totalBuyNum} 当前品种数量：{curHoldTypeNum}')
    
    buy_in_down_rate = context.Buy_In_Fall_Rate
    buy_in_up_rate = context.Buy_In_Up_Rate
    useClose = (context.Use_Close == 1)
    print(f'设置买入时（单只）下跌比例：{buy_in_down_rate}，买入时（单只）上涨比例：{buy_in_up_rate}， 是否使用昨收:{useClose}')
    
    avgBuyAmount = (totalCash / totalBuyNum) if (totalBuyNum > 0) else 0
    
    avgBuyAmount = ("%.2f" % avgBuyAmount)
    msg = f"设置的买入下跌比例为:{buy_in_down_rate} 买入总量:{totalBuyNum} 是否使用昨收:{useClose}"
    msg += f"\n资产总值:{totalCash} 目前已持仓种类数量:{curHoldTypeNum} 剩余现金:{leftCash}元 每股均分金额:{avgBuyAmount}元"
    log(msg)    

    # --------init中的测试函数，如果是循环执行的，应该加入到tick或者bar中
    
    '''
    pos = context.account().position(symbol = 'SHSE.605003', side = OrderSide_Buy)
    if not pos:
        curHolding = 0
        print(f'123 cur holding: 0')
    else:
        curHolding = pos.available_now
        #print(f'{tick.symbol} cur holding: {curHolding}')
        print(f"今持：{curHolding} 总持：{pos.volume} 可用：{pos.available_now}")
    '''


def reach_time(context, target_time):
    now = datetime.datetime.strptime(str(context.now.date()) + str(context.now.hour) + ":" + str(context.now.minute), '%Y-%m-%d%H:%M')
    target_time = datetime.datetime.strptime(str(context.now.date()) + target_time, '%Y-%m-%d%H:%M')

    return (now >= target_time)


def check_multi_up_flag(context, tick, rate, buy_in_up_rate):
    # 检测多次上涨后买入相关逻辑
    # 逻辑略微有点复杂，我们的多次上涨检测，需要检测是否超过目标值，超过的话，则计数count+=1
    # 且反转检测标记，开始检测是否低于目标值，低于以后，再次反转，准备检测下一次超过目标值，然后计数继续加1
    multi_up_buy_complete = True # 默认为已完成，在下列流程中修改（流程不开启的话，也是默认完成）
    
    if context.ids[tick.symbol].enable_multi_up_buy_flag:
        # 次数不等，就还需要进一步检测
        # 暂时用不到总次数了，注释记录一下，不过后面可能会用
        #if context.ids[tick.symbol].multi_up_cur_count != context.ids[tick.symbol].multi_up_total_count:
        reach_0945 = reach_time(context, context.ids[tick.symbol].multi_up_time_line)

        # 初始化（需要一个初始化检测）以及check标记为真的检查
        if context.ids[tick.symbol].multi_up_buy_need_check:
            if (rate > buy_in_up_rate) and (not reach_0945):
                # check标记为真的时候，rate高于目标后，计数加1，且转换check标记
                context.ids[tick.symbol].multi_up_buy_need_check = False
                context.ids[tick.symbol].multi_up_cur_count += 1
                log(f"[multi_up] tick.symbol reach {round(rate * 100, 3)}% with count{context.ids[tick.symbol].multi_up_cur_count}")
            else:
                # 需要check时，如果rate还低于目标，则什么都不做
                pass
        # 低于X%后的检测逻辑
        else:
            if (rate > buy_in_up_rate):
                # check标记反转后，表示已经高于目标值，这时如果继续高于，则说明都不做
                pass
            else:
                # check标记反转后，表示已经高于目标值，这时如果低于了，则应该把check标记为真，准备检测是否可以再次超过目标
                context.ids[tick.symbol].multi_up_buy_need_check = True

        # 在目标时间（比如09:45）之前，需要统计更新最高值
        if not reach_0945:
            open_fpr = (tick.price - tick.open) / tick.open
            if open_fpr > context.ids[tick.symbol].highest_open_fpr:
                context.ids[tick.symbol].highest_open_fpr = open_fpr
            # 时间线之前不进行购买，直接返回
            return False
        else:
            # 到达时间线后，目前的策略是，之前没有到过目标值的（count为0），我们才可以买入
            if context.ids[tick.symbol].multi_up_cur_count == 0:
                return True
            else:
                return False

    return multi_up_buy_complete


def check_buy_all_force_flag(context):

    unfixed_arr = {}
    buy_all_with_100 = 0
    for k,v in context.ids_buy_target_info_dict.items():
        if k not in context.ids_buy:
            continue
        # 只要还有无效数据，就不能开始计算分配，返回False
        if not context.ids_buy_target_info_dict[k].first_record_flag:
            if context.tick_count_for_buy_ok:
                print(f"未取得全部有效price，目前[{k}]还未进行过价格记录，返回False")
            return False
        # 已经分配好的情况（有任意一只分配好，即视为都分配好了，因为首先要通过每只都能买一手的验证）
        elif context.ids_buy_target_info_dict[k].fixed_buy_in_base_num > 0:
            return True
        
        # 因为存在停牌的特殊情况（涨停和跌停price好像不是0，是买卖5价格为0），我们需要自己跳过价格为0的标的，不处理它们
        if v.price == 0:
            continue
        
        unfixed_arr[k] = v.price
        buy_base_num = 1
        # 688开头的至少要买2手，需要特殊处理
        if ((".688" in k) or (".689" in k)):
            buy_base_num += 1

        buy_all_with_100 += (v.price * buy_base_num * 100.0)

    if len(unfixed_arr) == 0:
        return

    total_cash = context.account().cash['nav'] # 总资金
    total_cash = float('%.2f' % total_cash)
    left_cash = total_cash

    if (buy_all_with_100 > total_cash):
        print(f"[{context.now}]每只股买入1手的情况下，[{buy_all_with_100}]已经超过了总资金[{total_cash}]，策略无法实施！！")
        print(f"已经取消所有订阅，调整ID列表后，重新刷新")
        unsubscribe(symbols='*', frequency='tick')
        return False
    elif (buy_all_with_100 / total_cash) > 0.9:
        if context.tick_count_for_buy_ok:
            print(f"[警告]尝试买入所有目标的情况下，每个目标购买一手的情况下，占比超过总资金的90%！！很可能无法买入全部！！")

    avg_for_one = total_cash / len(context.ids_buy_target_info_dict) / 100.0

    fixed_arr = {}
    total_fixed_cash = 0

    need_handle_flag = True
    # 循环检测高于均值的对象，并剔除，同时记录剩余对象
    while need_handle_flag:
        left_arr = {}

        for k,v in unfixed_arr.items():
            if v > avg_for_one:

                buy_base_num = 1
                # 检测688的特殊情况
                if ((".688" in k) or (".689" in k)):
                    buy_base_num += 1

                fixed_arr[k] = v * buy_base_num
                total_fixed_cash += (v * buy_base_num * 100.0)
                left_cash -= (v * buy_base_num * 100.0)
                context.ids_buy_target_info_dict[k].fixed_buy_in_base_num = buy_base_num
            else:
                # 单价小于平均值的情况，也不能说明就直接归类到剩余了，688的x2后，其实也算是超过了
                if ((".688" in k) or (".689" in k)) and (v * 2.0 > avg_for_one):
                    buy_base_num = 2
                    fixed_arr[k] = v * buy_base_num
                    total_fixed_cash += (v * buy_base_num * 100.0)
                    left_cash -= (v * buy_base_num * 100.0)
                    context.ids_buy_target_info_dict[k].fixed_buy_in_base_num = buy_base_num
                else:
                    left_arr[k] = v

        # 先判断是否退出，当left_arr和unfixed_arr长度一样时，说明没有找到高于平均值的目标
        if len(left_arr) == len(unfixed_arr):
            print(f"已找到所有高于平均值的目标，进行后续分配计算")
            break

        # 没有退出while的话，说明需要计算下一轮，更新均值（根据剩余资金和剩余目标）
        avg_for_one = left_cash / len(left_arr) / 100.0
        # 更新剩余数组
        unfixed_arr = left_arr.copy()

    # 计算剩余数组中的对象，应该买入的基础数量
    avg_amount = avg_for_one * 100.0
    test_remainder = 10 # 循环测试余数，每次加10，找到一个临界值，各个不会超过剩余资金
    need_test_flag = True # 从10开始测试，一般都会超过剩余资金，直到找到合适的余数

    unfixed_total_use_cash = 0
    while need_test_flag:
        test_total_use_cash = 0

        for k,v in unfixed_arr.items():
            buy_in_num = math.floor(avg_amount / v)
            base_num = math.floor(buy_in_num / 100)
            remainder = buy_in_num % 100

            # 这个地方的处理还有待商榷，虽然之前测试情况可以买完，但是不一定适用所有情况，还需要思考
            if (remainder > test_remainder):
                base_num += 1
            
            context.ids_buy_target_info_dict[k].fixed_buy_in_base_num = base_num
            test_total_use_cash += (base_num * 100.0 * v)

        # 检查是否通过测试，没有超过总剩余资金的话，就是通过了
        if test_total_use_cash < left_cash:
            log(f"[statictics]余数检测成功，将使用余数[{test_remainder}]计算base_num，left_case:[{left_cash}] will_use[{test_total_use_cash}]")
            unfixed_total_use_cash = test_total_use_cash
            break
        
        # 没有通过测试的话，余数加10，余数越大，用到的剩余资金就会越少
        test_remainder += 10

    # 调试输出信息，不需要时可以关闭（不过这个只会输出一次，应该也没什么）
    fixed_all_equal_1 = True
    unfixed_all_greater_than_1 = True
    for k,v in fixed_arr.items():
        if context.ids_buy_target_info_dict[k].fixed_buy_in_base_num != 1:
            fixed_all_equal_1 = False
            break
    
    for k,v in unfixed_arr.items():
        if context.ids_buy_target_info_dict[k].fixed_buy_in_base_num < 1:
            unfixed_all_greater_than_1 = False
            break

    msg = f"[statistics]fixed_arr总共会使用cash为：[{total_fixed_cash}]， 总计使用cash:[{total_fixed_cash + unfixed_total_use_cash}]"
    msg += f"\nfixed_num[{len(fixed_arr)}], unfixed_num[{len(unfixed_arr)}], total_num[{len(unfixed_arr) + len(fixed_arr)}]"
    msg += f"\nfixed_arr all equal 1:[{fixed_all_equal_1}], unfixed_arr all greater than 1:[{unfixed_all_greater_than_1}]"
    log(msg)

    return True

# a1的脚本监控逻辑，监控a1的T-1到T-9的所有标的，记录超过3%的第一次时间，以及最后收盘时对比昨收的涨幅百分比
def monitor_center(context, tick):
    
    monitor_A1(context, tick)
    # 下面就可以加入其他监控的函数了，后面就从这里开始扩展
    # monitor_B(context, tick)
    # monitor_C(context, tick)

def monitor_A1(context, tick):
    # ------------------下面是一些openpyxl操作excel的常用函数
    # # 新建工作簿（不是打开先存的）
    # workbook = openpyxl.Workbook()
    # # 获取默认工作表
    # sheet = workbook.active
    # sheet.title = "title"

    # # 写入数据测试
    # sheet['A1'] = '姓名'
    # sheet['B1'] = '年龄'
    # # for row in sheet['A10:F12']:

    # # 添加一行数据
    # sheet.append(['Alice', 25])
    # # 对齐测试
    # align = Alignment(horizontal='right')
    # sheet.cell(4, 4).value = 'test'
    # sheet.cell(4, 4).alignment = align
    # # 设置字体颜色和背景色
    # fill = PatternFill('solid', fgColor='f8c600') #设置填充色
    # font = Font(u'黑体', size = 8, bold = True, italic = False, strike = False, color = '000000') # 设置字体样式
    # sheet.cell(4, 4).fill = fill
    # # sheet.cell(4, 4).font = font
    # # 设置单元格边框
    # thin = Side(border_style = 'thick', color = "ff0000") #设置边框为红色
    # border = Border(left = thin, right=thin, top = thin, bottom = thin)
    # sheet.cell(4, 4).border = border

    # # 合并单元格测试
    # sheet.merge_cells(start_row=1, end_row=1, start_column=3, end_column=4)

    # # 保存工作簿
    # workbook.save("c:/users/administrator/desktop/example.xlsx")

    if tick.symbol not in context.ids_buy_target_info_dict.keys():
        return
    
    t = 0
    if context.test_info == 2:
        t = time.time()
    
    curVal = context.ids_buy_target_info_dict[tick.symbol].price
    # 如果price不正常的话，则跳过对该标的的监控即可
    if curVal <= 0:
        return
    
    # 监控上涨幅度是否超过3%，是就进入记录池
    if context.ids_buy_target_info_dict[tick.symbol].pre_close == 0:
        info = get_instruments(symbols = tick.symbol, df = True)
        # empty情况一般就是ST的股票，直接先跳过不处理
        if info.empty:
            print(f"get info null, this should not happen......")
            return
            
        # 最好不要直接使用df的值，很多奇怪的现象，比如下面的float相除，如果插入了df数据，结果是对的，但是小数点只有2位。。。。
        # 一定要先从数组里拿出来
        context.ids_buy_target_info_dict[tick.symbol].pre_close = info.pre_close[0]
        context.ids_buy_target_info_dict[tick.symbol].name = info.sec_name[0]
        context.ids_buy_target_info_dict[tick.symbol].upper_limit = info.upper_limit[0]
        
    pre_close = context.ids_buy_target_info_dict[tick.symbol].pre_close
    # 掘金里的涨跌幅是需要自己手动计算的
    rate = (curVal - pre_close) / pre_close
    
    if context.test_info == 2:
        print(f"monitor phase 1 cost time:[{time.time() - t:.4f} s]")
    
    # 这里需要先初始化，而不是等到3%触发再初始化（会导致记录不到完整的最高和最低值）
    if (tick.symbol not in context.ids_a1_info_dict.keys()):
        context.ids_a1_info_dict[tick.symbol] = MonitorA1Info()
        
    # 超过3%就进行记录（只记录一次）
    if (rate >= 0.03) and (context.ids_a1_info_dict[tick.symbol].over_3_time == 0):
        context.ids_a1_info_dict[tick.symbol].over_3_time = context.now
        # 直接按顺序推入到记录的array里面，这样直接最后就是排序好的，不需要再次整理dict或者arr
        context.ids_a1_time_sorted_arr.append(tick.symbol)
        print(f"发现突破3%的标的[{tick.symbol}]")
        
    # 更新超过3%被记录dict中的最高和最低值（这里是错误的逻辑，不能再超过3%后再记录最高最低，这样的话，拿不到全天的最高和最低）
    if (rate > context.ids_a1_info_dict[tick.symbol].highest_rate):
        context.ids_a1_info_dict[tick.symbol].highest_rate = rate
        context.ids_a1_info_dict[tick.symbol].highest_rate_time = context.now
        print(f"正在更新[{tick.symbol}]高值到[{round(rate * 100, 2)}%]")
    # 更新最低值不能写成elif，这个错误犯了超过2次了，这里备注一下，最高和最低需要同时统计，否则一直在涨的话，记录的最低值就是有bug的（比如从0%到9%一直涨，我们都记录不了最低值，其实最低应该是0%，但是最后可能就记录到8.9%这样的数据）
    if (rate < context.ids_a1_info_dict[tick.symbol].lowest_rate):
        context.ids_a1_info_dict[tick.symbol].lowest_rate = rate
        context.ids_a1_info_dict[tick.symbol].lowest_rate_time = context.now
        print(f"正在更新[{tick.symbol}]低值到[{round(rate * 100, 2)}%]")
        
    if context.test_info == 2:
        print(f"monitor phase 2 cost time:[{time.time() - t:.4f} s]")
        
    reach_1530 = reach_time(context, "15:03")
    # 到达15点30后，开始统计并输出今天的结果，生成excel文件，只输出一次
    if ((not context.ids_a1_excel_generated_flag) and (reach_1530)) or (context.ids_a1_manual_generate_excel_flag):
        # 转换标识，只自动输出一次文件，且非手动激活的情况下
        if (not context.ids_a1_manual_generate_excel_flag):
            context.ids_a1_excel_generated_flag = True
        context.ids_a1_manual_generate_excel_flag = False
        print(f"开始生成Excel文件--------")
        
        # 生成所有记录超过3%的标的的收盘涨幅
        for k,v in context.ids_a1_info_dict.items():
            if v.over_3_time == 0:
                continue
            pre_close = context.ids_buy_target_info_dict[k].pre_close
            cur_price = context.ids_buy_target_info_dict[k].price
            close_rate = (cur_price - pre_close) / pre_close
            context.ids_a1_info_dict[k].close_rate = close_rate
            
        # 对收盘涨跌幅进行一次排序
        # 对dict先进行排序
        # 已经测试过了，x: x[1]的意思就是items中的索引1（也就是key(0)，value(1)中的value），如果value是数组的话，还可以进一步指定二维数组的位置
        dict_a1_for_sort = {}
        for k,v in context.ids_a1_info_dict.items():
            if v.over_3_time == 0:
                continue
            dict_a1_for_sort[k] = v.close_rate
        sorted_a1_close_rate_dict = dict(sorted(dict_a1_for_sort.items(), key=lambda x: x[1], reverse=False))
        sorted_a1_close_rate_arr = []
        for k in sorted_a1_close_rate_dict.keys():
            sorted_a1_close_rate_arr.append(k)
            
        # 根据最终数据，开始生成Excel文件--------
        # 新建工作簿（不是打开先存的）
        workbook = openpyxl.Workbook()
        # 获取默认工作表
        sheet = workbook.active
        
        # 对齐设置
        align_right = Alignment(horizontal='right')
        align_center = Alignment(horizontal='center')
        align_left = Alignment(horizontal='left')
        
        # 填充颜色设置
        fill_color1 = PatternFill('solid', fgColor='f8c600')
        fill_color2 = PatternFill('solid', fgColor='f80000')

        # 写入数据测试
        sheet.cell(2,1).value = '代码'
        sheet.cell(2,2).value = '名称'
        sheet.cell(2,3).value = '当日最低'
        sheet.cell(2,4).value = '触发时间'
        sheet.cell(2,5).value = '触发3%时间'
        sheet.cell(2,6).value = '当日最高'
        sheet.cell(2,7).value = '触发时间'
        sheet.cell(2,8).value = '收盘幅度'

        # 合并单元格测试
        sheet.merge_cells(start_row=1, end_row=1, start_column=9, end_column=12)
        sheet.merge_cells(start_row=1, end_row=1, start_column=13, end_column=16)

        sheet.cell(1,9).value = '按照3%触发排序'
        sheet.cell(1,9).alignment = align_center
        sheet.cell(1,9).fill = fill_color1
        sheet.cell(2,9).value = '代码'
        sheet.cell(2,10).value = '名称'
        sheet.cell(2,11).value = '3%触发'
        sheet.cell(2,12).value = '收盘幅度'

        sheet.cell(1,13).value = '按照收盘幅度排序' # M,N,O,P
        sheet.cell(1,13).alignment = align_center
        sheet.cell(1,13).fill = fill_color2
        sheet.cell(2,13).value = '代码'
        sheet.cell(2,14).value = '名称'
        sheet.cell(2,15).value = '%3触发'
        sheet.cell(2,16).value = '收盘幅度'
        
        # 正式的数据从第三行开始写（由于表格的固定设置）
        start_row = 3
        for k,v in context.ids_a1_info_dict.items():
            if v.over_3_time == 0:
                continue
            # 写入excel一共分3个批次：
            # 1.当日的最高最低以及附带数据
            # 2.按照3%触发顺序的一个表格
            # 3.按照当日收盘幅度的一个表格
            sheet.cell(start_row, 1).value = k[5:]
            sheet.cell(start_row, 2).value = context.ids_buy_target_info_dict[k].name
            sheet.cell(start_row, 3).value = str(round(context.ids_a1_info_dict[k].lowest_rate * 100, 2)) + "%"
            sheet.cell(start_row, 4).value = str(context.ids_a1_info_dict[k].lowest_rate_time.hour) + ":" + str(context.ids_a1_info_dict[k].lowest_rate_time.minute)
            sheet.cell(start_row, 5).value = str(context.ids_a1_info_dict[k].over_3_time.hour) + ":" + str(context.ids_a1_info_dict[k].over_3_time.minute)
            sheet.cell(start_row, 6).value = str(round(context.ids_a1_info_dict[k].highest_rate * 100, 2)) + "%"
            sheet.cell(start_row, 7).value = str(context.ids_a1_info_dict[k].highest_rate_time.hour) + ":" + str(context.ids_a1_info_dict[k].highest_rate_time.minute)
            sheet.cell(start_row, 8).value = str(round(context.ids_a1_info_dict[k].close_rate * 100, 2)) + "%"
            
            key_for_3_rate = context.ids_a1_time_sorted_arr[start_row - 3]
            sheet.cell(start_row, 9).value = key_for_3_rate[5:]
            sheet.cell(start_row, 10).value = context.ids_buy_target_info_dict[key_for_3_rate].name
            sheet.cell(start_row, 11).value = str(context.ids_a1_info_dict[key_for_3_rate].over_3_time.hour) + ":" + str(context.ids_a1_info_dict[key_for_3_rate].over_3_time.minute)
            sheet.cell(start_row, 12).value = str(round(context.ids_a1_info_dict[key_for_3_rate].close_rate * 100, 2)) + "%"
            
            key_for_close_rate = sorted_a1_close_rate_arr[start_row - 3]
            sheet.cell(start_row, 13).value = key_for_close_rate[5:]
            sheet.cell(start_row, 14).value = context.ids_buy_target_info_dict[key_for_close_rate].name
            sheet.cell(start_row, 15).value = str(context.ids_a1_info_dict[key_for_close_rate].over_3_time.hour) + ":" + str(context.ids_a1_info_dict[key_for_close_rate].over_3_time.minute)
            sheet.cell(start_row, 16).value = str(round(context.ids_a1_info_dict[key_for_close_rate].close_rate * 100, 2)) + "%"
            
            # 每写完一行，移动一次工作行数
            start_row += 1
            

        # 保存工作簿
        workbook.save(f"c:/users/administrator/desktop/A1监控-{context.now.date()}.xlsx")
        print(f"生成Excel文件结束--------")
        
    if context.test_info == 2:
        print(f"monitor phase 3 cost time:[{time.time() - t:.4f} s]")


def on_bar(context, bars):
    print('---------on_bar---------')
    # print(bars)


def on_tick(context, tick):
    if context.tick_count_for_statistics == 0:        
        context.record_time = time.time()
        
    context.tick_count_for_statistics += 1
    if context.test_info == 1:
        print(f'---------on_tick({tick.symbol})---------')
    # #print(tick)
    # info = get_instruments(symbols = tick.symbol, df = True)
    # msg = f"\n-------->>开始尝试卖出[{info.sec_id[0]}:{info.sec_name[0]}]"
    # print(msg)
    # return
    
    t = time.time()

    if context.ids[tick.symbol].buy_flag == 1:
        if (tick.price > 0) and (tick.symbol in context.ids_buy_target_info_dict.keys()):
            context.ids_buy_target_info_dict[tick.symbol].price = tick.price
            context.ids_buy_target_info_dict[tick.symbol].first_record_flag = True
        # else:
        #     print(f"potential [buy]price error, == 0, [{tick.symbol}]")
    # if context.ids[tick.symbol].buy_flag == 0:
    #     # 为什么这里判断大于0，因为这个订阅很扯，9点30之前有些票会给你发价格0过来，3点以后有些之前有效的票，也会发0给你，所以必须判断
    #     if tick.price > 0:
    #         context.ids_virtual_sell_target_info_dict[tick.symbol].price = tick.price
    #         context.ids_virtual_sell_target_info_dict[tick.symbol].first_record_flag = True

    # 根据设置的tick handle frequence来跳过tick的处理
    # context.ids[tick.symbol].tick_cur_count += 1
    # if context.ids[tick.symbol].tick_cur_count < context.ids[tick.symbol].tick_handle_frequence:
    #     return
    # else:
    #     context.ids[tick.symbol].tick_cur_count = 0

    # 检测是否已经记录所有的buy或者sell的price，这样后续流程就不用再浪费性能检测了，该过程激活一次就ok了（要注意refresh中重置flag，因为刷新可能ids变更）
    invalid_buy_symbol = ""
    if (not context.get_all_buy_price_flag):
        record_all = True
        for k,v in context.ids_buy_target_info_dict.items():
            if v.price == 0:
                record_all = False
                invalid_buy_symbol = k
                break
        if record_all:
            context.get_all_buy_price_flag = True

    # invalid_sell_symbol = ""
    # if (not context.get_all_sell_price_flag):
    #     record_all = True
    #     for k,v in context.ids_virtual_sell_target_info_dict.items():
    #         # 这里用过first_record_flag比较，但是不行，验证发现price第一次记录了也很可能是0，原因不明
    #         # 使用还是必须要用price，否则就会出现之前刚刚开盘统计数据最低就到-4，-5，甚至更低，就是因为好几只票，明明有价格，但是第一次进来给的0
    #         if v.price == 0:
    #             invalid_sell_symbol = k
    #             record_all = False
    #             break
    #     if record_all:
    #         context.get_all_sell_price_flag = True

    # 更新缓存价格，在外面更新，里面return的条件太多了，这样更新的价格，statistics也可以用
    if context.tick_count_for_statistics >= len(context.ids):
        context.tick_count_for_statistics = 0
        print(f"目前整个监控全部循环({len(context.ids)})一次耗时: {time.time() - context.record_time:.4f} s")
        if invalid_buy_symbol != "":
            print(f"卖出列表中存在无法获取price的情况[{invalid_buy_symbol}]，等待一段时间（30s）后仍然无法获取到的，说明该标的存在某些异常，但是对于监控脚本可以不进行处理")
        if context.test_info == 1:
            print(f"trade logic cost time with count limit------: {time.time() - t:.4f} s")

    now = datetime.datetime.strptime(str(context.now.date()) + str(context.now.hour) + ":" + str(context.now.minute), '%Y-%m-%d%H:%M')
    target_time = datetime.datetime.strptime(str(context.now.date()) + context.start_operation_time, '%Y-%m-%d%H:%M')
    if (now < target_time):
        # 没有到9点30，先不要开始操作（目前原因，是我统计剩余ids_buy长度，有些股我看了竞价就可以买，但是不会成功，会被我超时取消掉）
        # 现在先简单处理，9点30前都不操作，后面有需要再调整
        print(f"not reach the begin time [{context.start_operation_time}]......")
        return

    if context.get_all_buy_price_flag:
        context.tick_count_for_buy += 1
        if (context.tick_count_for_buy > (len(context.ids) * context.tick_count_limit_rate_for_buy)):
            context.tick_count_for_buy_ok = True
        else:
            context.tick_count_for_buy_ok = False
    # if context.get_all_sell_price_flag:
    #     context.tick_count_for_sell += 1
    #     if (context.tick_count_for_sell > (len(context.ids) * context.tick_count_limit_rate_for_sell)):
    #         context.tick_count_for_sell_ok = True
    #     else:
    #         context.tick_count_for_sell_ok = False
            
    # if (not context.get_all_buy_price_flag):
    #     print(f"目前尚未获得所有")

    # 根据不同策略尝试买卖--------
    # 策略B（对冲整体买卖策略，不存在单只处理）
    if (context.strategy_info.M == 1) and (context.ids[tick.symbol].buy_flag == 1):
        monitor_center(context, tick)
    else:
        print(f"[error] enter trade strategy failed, this should never happen")

    if context.get_all_buy_price_flag:
        if (context.tick_count_for_buy > (len(context.ids) * context.tick_count_limit_rate_for_buy)):
            context.tick_count_for_buy = 0
    if context.get_all_sell_price_flag:
        if (context.tick_count_for_sell > (len(context.ids) * context.tick_count_limit_rate_for_sell)):
            context.tick_count_for_sell = 0 # 重置tick count

    if context.test_info == 1:
        print(f"trade logic cost time: {time.time() - t:.4f} s")


# 委托执行回报事件
# 响应委托被执行事件，委托成交或者撤单拒绝后被触发。
# 注意：
# 1、交易账户重连后，会重新推送一遍交易账户登录成功后查询回来的所有执行回报
# 2、撤单拒绝后，会推送撤单拒绝执行回报，可以根据exec_id区分
def on_execution_report(context, execrpt):
    print('--------on_execution_report')
    #print(execrpt)

# 交易账户状态更新事件
# 响应交易账户状态更新事件，交易账户状态变化时被触发。
# 交易账户状态对象，仅响应 已连接，已登录，已断开 和 错误 事件
def on_account_status(context, account):
    #print('--------on_account_status')
    #print(account)
    
    if (account['status']['state'] != 3):
        log(f'账号连接异常！！！目前没有任何处理')

    if (account['status']['error']['code'] != 0):
        log(f"发生账号错误：{account['status']['error']['code']}！！！目前没有任何处理")

if __name__ == '__main__':
    '''
        strategy_id策略ID, 由系统生成
        filename文件名, 请与本文件名保持一致
        mode运行模式, 实时模式:MODE_LIVE回测模式:MODE_BACKTEST
        token绑定计算机的ID, 可在系统设置-密钥管理中生成
        backtest_start_time回测开始时间
        backtest_end_time回测结束时间
        backtest_adjust股票复权方式, 不复权:ADJUST_NONE前复权:ADJUST_PREV后复权:ADJUST_POST
        backtest_initial_cash回测初始资金
        backtest_commission_ratio回测佣金比例
        backtest_slippage_ratio回测滑点比例
        '''
    run(strategy_id='1457c400-f485-11ed-a1a9-005056bc063c',
        filename='main.py',
        #mode=MODE_BACKTEST,
        mode=MODE_LIVE,
        token='4f0478a8560615e1a0049e2e2565955620b3ec02',
        backtest_start_time='2020-11-01 08:00:00',
        backtest_end_time='2020-11-10 16:00:00',
        backtest_adjust=ADJUST_PREV,
        backtest_initial_cash=10000000,
        backtest_commission_ratio=0.0001,
        backtest_slippage_ratio=0.0001)
